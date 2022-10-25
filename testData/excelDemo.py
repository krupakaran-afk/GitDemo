import openpyxl

book = openpyxl.load_workbook("C:\\Users\\krupa\\Downloads\\PythonDemo.xlsx")
sheet = book.active
Dict = {}
# read
cell = sheet.cell(row=1, column=2)
print(cell.value)
# write
sheet.cell(row=2, column=2).value = "krupa"

sheet.max_row
sheet.max_column

sheet['A2'].value

for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == "Testcase2":
        for j in range(2, sheet.max_column + 1):
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        # print(sheet.cell(row=i, column=j).value)
print(Dict)
